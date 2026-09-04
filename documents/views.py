from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Count, F
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.conf import settings
import json, datetime, csv, os

from .models import (Document, Category, Notification,
                     InnovationProject, Training, UserProfile, ActivityLog)
from .forms import DocumentForm, DocumentSearchForm, CategoryForm
from .activity import log_activity, notify_users
from .decorators import registrar_required, admin_required
from .utils import generate_reference_number, generate_qr_code


# ──────────────────────────── AUTH ───────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        login(request, user)
        messages.success(request, f'እንኳን ደህና መጡ, {user.get_full_name() or user.username}!')
        return redirect(request.GET.get('next', 'dashboard'))
    return render(request, 'documents/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


# ──────────────────────────── DASHBOARD ──────────────────────────────

@login_required
def dashboard(request):
    today      = timezone.localdate()
    week_start = today - datetime.timedelta(days=today.weekday())
    docs       = Document.objects.all()

    total       = docs.count()
    incoming    = docs.filter(doc_type='INCOMING').count()
    outgoing    = docs.filter(doc_type='OUTGOING').count()
    internal    = docs.filter(doc_type='INTERNAL').count()
    pending     = docs.filter(status='PENDING').count()
    in_progress = docs.filter(status='IN_PROGRESS').count()
    completed   = docs.filter(status='COMPLETED').count()
    urgent      = docs.filter(priority='URGENT').count()
    overdue     = docs.filter(due_date__lt=today).exclude(
                    status__in=['COMPLETED', 'ARCHIVED']).count()
    today_docs  = docs.filter(created_at__date=today).count()
    week_docs   = docs.filter(created_at__date__gte=week_start).count()
    my_docs     = docs.filter(assigned_to=request.user).exclude(
                    status__in=['COMPLETED', 'ARCHIVED']).count()

    recent_docs   = docs.select_related('category', 'created_by').order_by('-created_at')[:8]
    deadline_soon = docs.filter(
        due_date__gte=today,
        due_date__lte=today + datetime.timedelta(days=7)
    ).exclude(status__in=['COMPLETED', 'ARCHIVED']).order_by('due_date')[:6]

    chart_labels, chart_data = [], []
    for i in range(5, -1, -1):
        d = today - datetime.timedelta(days=30 * i)
        chart_labels.append(d.strftime('%b %Y'))
        chart_data.append(docs.filter(created_at__year=d.year, created_at__month=d.month).count())

    type_data    = {'ገቢ ደብዳቤ': incoming, 'ወጪ ደብዳቤ': outgoing, 'የ ማስታወሻ': internal}
    status_labels = [s[1] for s in Document.STATUS_CHOICES]
    status_values = [docs.filter(status=s[0]).count() for s in Document.STATUS_CHOICES]

    return render(request, 'documents/dashboard.html', {
        'total': total, 'incoming': incoming, 'outgoing': outgoing, 'internal': internal,
        'pending': pending, 'in_progress': in_progress, 'completed': completed,
        'urgent': urgent, 'overdue': overdue,
        'today_docs': today_docs, 'week_docs': week_docs, 'my_docs': my_docs,
        'recent_docs': recent_docs, 'deadline_soon': deadline_soon,
        'chart_labels':  json.dumps(chart_labels),
        'chart_data':    json.dumps(chart_data),
        'type_labels':   json.dumps(list(type_data.keys())),
        'type_values':   json.dumps(list(type_data.values())),
        'status_labels': json.dumps(status_labels),
        'status_values': json.dumps(status_values),
        'today': today,
        'completion_pct': int((completed / total * 100) if total else 0),
    })


# ──────────────────────────── DOCUMENT LIST ──────────────────────────

@login_required
def document_list(request):
    form = DocumentSearchForm(request.GET or None)
    docs = Document.objects.select_related('category', 'created_by').all()

    if form.is_valid():
        q = form.cleaned_data.get('query')
        if q:
            docs = docs.filter(
                Q(title__icontains=q) | Q(reference_number__icontains=q) |
                Q(sender__icontains=q) | Q(receiver__icontains=q) |
                Q(description__icontains=q)
            )
        for field in ('doc_type', 'status', 'priority', 'category'):
            val = form.cleaned_data.get(field)
            if val:
                docs = docs.filter(**{field: val})
        if form.cleaned_data.get('date_from'):
            docs = docs.filter(created_at__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            docs = docs.filter(created_at__date__lte=form.cleaned_data['date_to'])

    if request.GET.get('export') == 'csv':
        return _export_csv(docs)

    # Sorting
    sort = request.GET.get('sort', '-created_at')
    allowed_sorts = ['reference_number', '-reference_number', 'created_at', '-created_at',
                     'due_date', '-due_date', 'title', '-title']
    if sort not in allowed_sorts:
        sort = '-created_at'
    docs = docs.order_by(sort)

    paginator = Paginator(docs, 15)
    page_obj  = paginator.get_page(request.GET.get('page'))
    today     = timezone.localdate()
    return render(request, 'documents/document_list.html', {
        'form': form, 'page_obj': page_obj,
        'total_results': paginator.count,
        'today_date': today,
        'today_plus7': today + datetime.timedelta(days=7),
    })


def _export_csv(queryset):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="documents.csv"'
    response.write('\ufeff')
    w = csv.writer(response)
    w.writerow(['ቁጥር', 'ርዕስ', 'ዓይነት', 'ሁኔታ', 'ቅድሚያ', 'ላኪ', 'ተቀባይ', 'ቀን'])
    for d in queryset:
        w.writerow([
            d.reference_number, d.title, d.get_doc_type_display(),
            d.get_status_display(), d.get_priority_display(),
            d.sender, d.receiver, d.created_at.strftime('%d/%m/%Y'),
        ])
    return response


# ──────────────────────────── DOCUMENT DETAIL ────────────────────────

@login_required
def document_detail(request, pk):
    doc        = get_object_or_404(Document, pk=pk)
    activities = doc.activities.select_related('user').order_by('-timestamp')[:20]
    today      = timezone.localdate()
    is_overdue = (doc.due_date and doc.due_date < today
                  and doc.status not in ['COMPLETED', 'ARCHIVED'])
    Document.objects.filter(pk=pk).update(view_count=F('view_count') + 1)
    return render(request, 'documents/document_detail.html', {
        'doc': doc, 'activities': activities, 'is_overdue': is_overdue,
    })


# ──────────────────────────── CREATE ─────────────────────────────────

@login_required
@registrar_required
def document_create(request):
    form = DocumentForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        doc = form.save(commit=False)
        doc.created_by = request.user
        doc.save()
        log_activity(doc, request.user, 'ደብዳቤ ተመዝግቧል', f'"{doc.title}" ተመዝግቧል።')
        notify_users(
            f'አዲስ ደብዳቤ "{doc.reference_number} — {doc.title}" ተመዝግቧል።',
            link=f'/documents/{doc.pk}/', exclude_user=request.user,
        )
        messages.success(request, f'ደብዳቤ "{doc.title}" ተመዝግቧል!')
        return redirect('document_detail', pk=doc.pk)
    return render(request, 'documents/document_form.html', {'form': form, 'action': 'ምዝገባ'})


# ──────────────────────────── UPDATE ─────────────────────────────────

@login_required
@admin_required
def document_update(request, pk):
    doc        = get_object_or_404(Document, pk=pk)
    old_status = doc.status
    form       = DocumentForm(request.POST or None, request.FILES or None, instance=doc)
    if request.method == 'POST' and form.is_valid():
        updated = form.save()
        if old_status != updated.status:
            log_activity(doc, request.user, 'ሁኔታ ተቀይሯል',
                         f'"{doc.get_status_display()}" → "{updated.get_status_display()}"')
        else:
            log_activity(doc, request.user, 'ደብዳቤ ተዘምኗል')
        messages.success(request, 'ደብዳቤ ተዘምኗል!')
        return redirect('document_detail', pk=doc.pk)
    return render(request, 'documents/document_form.html', {
        'form': form, 'doc': doc, 'action': 'ማስተካከያ'
    })


# ──────────────────────────── DELETE ─────────────────────────────────

@login_required
@admin_required
def document_delete(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
        title = doc.title
        doc.delete()
        messages.warning(request, f'ደብዳቤ "{title}" ተሰርዟል።')
        return redirect('document_list')
    return render(request, 'documents/document_confirm_delete.html', {'doc': doc})


# ──────────────────────────── QUICK STATUS ───────────────────────────

@login_required
@admin_required
@require_POST
def update_status(request, pk):
    doc        = get_object_or_404(Document, pk=pk)
    new_status = request.POST.get('status')
    valid      = [s[0] for s in Document.STATUS_CHOICES]
    if new_status in valid:
        old = doc.get_status_display()
        doc.status = new_status
        doc.save(update_fields=['status', 'updated_at'])
        log_activity(doc, request.user, 'ሁኔታ ተቀይሯል',
                     f'"{old}" → "{doc.get_status_display()}"')
        return JsonResponse({'success': True, 'new_status': doc.get_status_display()})
    return JsonResponse({'success': False}, status=400)


# ──────────────────────────── DOWNLOAD ───────────────────────────────

@login_required
def document_download(request, pk):
    from django.http import FileResponse, Http404
    doc = get_object_or_404(Document, pk=pk)
    if not doc.file:
        messages.error(request, 'ፋይሉ አልተገኘም።')
        return redirect('document_detail', pk=pk)
    try:
        file_path = doc.file.path
        if not os.path.exists(file_path):
            raise Http404
        Document.objects.filter(pk=pk).update(download_count=F('download_count') + 1)
        log_activity(doc, request.user, 'ፋይል ወረደ',
                     f'"{os.path.basename(file_path)}" ወረደ።')
        response = FileResponse(open(file_path, 'rb'), as_attachment=True)
        response['Content-Disposition'] = (
            f'attachment; filename="{os.path.basename(file_path)}"'
        )
        return response
    except Exception:
        messages.error(request, 'ፋይሉ ከሰርቨሩ ላይ አልተገኘም።')
        return redirect('document_detail', pk=pk)


# ──────────────────────────── PRINT & QR ─────────────────────────────

@login_required
def document_print(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    return render(request, 'documents/document_print.html', {'doc': doc})


@login_required
def document_qr(request, pk):
    """Return the QR code PNG — with Woldiya logo embedded in the centre."""
    doc      = get_object_or_404(Document, pk=pk)
    scheme   = 'https' if request.is_secure() else 'http'
    doc_url  = f"{scheme}://{request.get_host()}/documents/{doc.pk}/"
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'woldiya_logo.png')
    png_bytes = generate_qr_code(doc_url, logo_path=logo_path)
    return HttpResponse(png_bytes, content_type='image/png')


# ──────────────────────────── OVERDUE ────────────────────────────────

@login_required
def overdue_documents(request):
    today = timezone.localdate()
    docs  = Document.objects.filter(due_date__lt=today).exclude(
                status__in=['COMPLETED', 'ARCHIVED']).order_by('due_date')
    return render(request, 'documents/overdue.html', {'docs': docs, 'today': today})


# ──────────────────────────── NOTIFICATIONS ──────────────────────────

@login_required
def notifications_view(request):
    # Mark all as read on visit
    if request.GET.get('mark_all'):
        request.user.notifications.filter(is_read=False).update(is_read=True)
        messages.success(request, 'ሁሉም ማሳወቂያዎች ተነቡ።')
        return redirect('notifications')

    notifs = request.user.notifications.order_by('-created_at')[:100]
    # Group by date
    from itertools import groupby
    from django.utils.timezone import localdate
    grouped = {}
    for n in notifs:
        day = localdate(n.created_at)
        grouped.setdefault(day, []).append(n)

    return render(request, 'documents/notifications.html', {
        'notifs': notifs,
        'grouped': grouped,
        'unread_count': request.user.notifications.filter(is_read=False).count(),
        'today': timezone.localdate(),
    })


@login_required
@require_POST
def mark_notification_read(request, pk):
    n = get_object_or_404(Notification, pk=pk, user=request.user)
    n.is_read = True
    n.save(update_fields=['is_read'])
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    return redirect(n.link or 'notifications')


@login_required
@require_POST
def mark_all_notifications_read(request):
    request.user.notifications.filter(is_read=False).update(is_read=True)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    messages.success(request, 'ሁሉም ማሳወቂያዎች ተነቡ።')
    return redirect('notifications')


# ──────────────────────────── AUDIT TRAIL ────────────────────────────

@login_required
def audit_trail(request):
    logs = ActivityLog.objects.select_related('document', 'user').order_by('-timestamp')

    # Filters
    user_q   = request.GET.get('user', '').strip()
    action_q = request.GET.get('action', '').strip()
    date_f   = request.GET.get('date_from', '').strip()
    date_t   = request.GET.get('date_to', '').strip()

    if user_q:
        logs = logs.filter(
            Q(user__username__icontains=user_q) |
            Q(user__first_name__icontains=user_q) |
            Q(user__last_name__icontains=user_q)
        )
    if action_q:
        logs = logs.filter(action__icontains=action_q)
    if date_f:
        try:
            logs = logs.filter(timestamp__date__gte=datetime.date.fromisoformat(date_f))
        except ValueError:
            pass
    if date_t:
        try:
            logs = logs.filter(timestamp__date__lte=datetime.date.fromisoformat(date_t))
        except ValueError:
            pass

    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="audit_trail.csv"'
        response.write('\ufeff')
        w = csv.writer(response)
        w.writerow(['ቀን', 'ተጠቃሚ', 'ደብዳቤ', 'እርምጃ', 'ዝርዝር'])
        for log in logs:
            w.writerow([
                log.timestamp.strftime('%d/%m/%Y %H:%M'),
                log.user.get_full_name() or log.user.username if log.user else '—',
                log.document.reference_number,
                log.action,
                log.detail or '—',
            ])
        return response

    paginator = Paginator(logs, 30)
    page_obj  = paginator.get_page(request.GET.get('page'))
    users     = User.objects.filter(is_active=True).order_by('username')
    return render(request, 'documents/audit_trail.html', {
        'page_obj': page_obj, 'users': users,
        'filter_user': user_q, 'filter_action': action_q,
        'filter_date_from': date_f, 'filter_date_to': date_t,
    })


# ──────────────────────────── REPORTS ────────────────────────────────

@login_required
def reports(request):
    today = timezone.localdate()
    docs  = Document.objects.all()

    by_type     = {t: docs.filter(doc_type=t).count() for t, _ in Document.DOCUMENT_TYPES}
    by_status   = {s: docs.filter(status=s).count()   for s, _ in Document.STATUS_CHOICES}
    by_priority = {p: docs.filter(priority=p).count() for p, _ in Document.PRIORITY_CHOICES}
    by_category = Category.objects.annotate(cnt=Count('documents')).order_by('-cnt')

    monthly = []
    for i in range(11, -1, -1):
        d = today.replace(day=1) - datetime.timedelta(days=30 * i)
        monthly.append({
            'label': d.strftime('%b %Y'),
            'count': docs.filter(created_at__year=d.year, created_at__month=d.month).count(),
        })

    if request.GET.get('export') == 'excel':
        return _export_excel_report(docs, by_type, by_status, by_priority, monthly, today)
    if request.GET.get('export') == 'pdf':
        return _export_pdf_report(docs, by_type, by_status, by_priority, monthly, today)

    return render(request, 'documents/reports.html', {
        'total': docs.count(),
        'by_type': by_type, 'by_status': by_status,
        'by_priority': by_priority, 'by_category': by_category,
        'monthly': monthly, 'today': today,
    })


def _export_excel_report(docs, by_type, by_status, by_priority, monthly, today):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl አልተጫነም። pip install openpyxl', status=500)

    wb = openpyxl.Workbook()
    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='1B4F72')
    center    = Alignment(horizontal='center')

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = 'ማጠቃለያ'
    ws.append(['ወልድያ ከተማ አስተዳደር — EDMS ሪፖርት'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'ዛሬ: {today}'])
    ws.append([])
    ws.append(['ጠቅላላ', 'ዓይነት', 'ብዛት'])
    for cell in ws[4]:
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
    ws.append(['ጠቅላላ ደብዳቤ', '', docs.count()])
    for k, v in by_type.items():
        ws.append([k, '', v])
    ws.append([])
    ws.append(['ሁኔታ', '', 'ብዛት'])
    for k, v in by_status.items():
        ws.append([k, '', v])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['C'].width = 12

    # ── Sheet 2: Monthly ──
    ws2 = wb.create_sheet('ወርሃዊ')
    ws2.append(['ወር', 'ደብዳቤ ብዛት'])
    for cell in ws2[1]:
        cell.font = hdr_font; cell.fill = hdr_fill
    for row in monthly:
        ws2.append([row['label'], row['count']])
    ws2.column_dimensions['A'].width = 18

    # ── Sheet 3: All documents ──
    ws3 = wb.create_sheet('ሁሉም ደብዳቤዎች')
    ws3.append(['ቁጥር', 'ርዕስ', 'ዓይነት', 'ሁኔታ', 'ቅድሚያ', 'ላኪ', 'ተቀባይ', 'ቀን'])
    for cell in ws3[1]:
        cell.font = hdr_font; cell.fill = hdr_fill
    for d in docs.order_by('-created_at'):
        ws3.append([
            d.reference_number, d.title, d.get_doc_type_display(),
            d.get_status_display(), d.get_priority_display(),
            d.sender, d.receiver, d.created_at.strftime('%d/%m/%Y'),
        ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws3.column_dimensions[col].width = 20

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="edms_report_{today}.xlsx"'
    return response


def _export_pdf_report(docs, by_type, by_status, by_priority, monthly, today):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return HttpResponse('reportlab አልተጫነም። pip install reportlab', status=500)

    from io import BytesIO
    buf  = BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                              rightMargin=2*cm, leftMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    brand  = colors.HexColor('#1B4F72')
    gold   = colors.HexColor('#C8960C')
    story  = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                  textColor=brand, fontSize=18, spaceAfter=6)
    story.append(Paragraph('ወልድያ ከተማ አስተዳደር — EDMS ሪፖርት', title_style))
    story.append(Paragraph(f'ዛሬ: {today}', styles['Normal']))
    story.append(HRFlowable(width='100%', color=gold, thickness=2, spaceAfter=12))

    # Summary table
    summary_data = [['ምድብ', 'ብዛት']]
    summary_data.append(['ጠቅላላ ደብዳቤ', str(docs.count())])
    for k, v in by_type.items():
        summary_data.append([k, str(v)])
    for k, v in by_status.items():
        summary_data.append([k, str(v)])

    t = Table(summary_data, colWidths=[12*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), brand),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN',      (1, 0), (1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('PADDING',    (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    # Monthly table
    story.append(Paragraph('ወርሃዊ ምዝገባ', styles['Heading2']))
    month_data = [['ወር', 'ደብዳቤ']]
    for row in monthly:
        month_data.append([row['label'], str(row['count'])])
    t2 = Table(month_data, colWidths=[12*cm, 4*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), brand),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN',      (1, 0), (1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('PADDING',    (0, 0), (-1, -1), 6),
    ]))
    story.append(t2)

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="edms_report_{today}.pdf"'
    return response


# ──────────────────────────── CAPACITY BUILDING ──────────────────────

@login_required
def capacity_building(request):
    projects  = InnovationProject.objects.select_related('lead').order_by('-created_at')
    trainings = Training.objects.prefetch_related('participants').order_by('-date')
    return render(request, 'documents/capacity_building.html', {
        'projects': projects, 'trainings': trainings,
    })


# ──────────────────────────── USER MANAGEMENT ────────────────────────

@login_required
def user_management(request):
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'ፈቃድ የለዎትም።')
        return redirect('dashboard')
    users = User.objects.select_related('profile').filter(is_active=True).order_by('username')
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        role    = request.POST.get('role')
        dept    = request.POST.get('department', '').strip()
        try:
            profile = UserProfile.objects.get(user_id=user_id)
            if role in dict(UserProfile.ROLE_CHOICES):
                profile.role       = role
                profile.department = dept
                profile.save()
                messages.success(request, 'ተጠቃሚ ተዘምኗል።')
        except UserProfile.DoesNotExist:
            messages.error(request, 'ተጠቃሚ አልተገኘም።')
        return redirect('user_management')
    return render(request, 'documents/user_management.html', {
        'users': users, 'role_choices': UserProfile.ROLE_CHOICES,
    })


# ──────────────────────────── API HELPERS ────────────────────────────

@login_required
def suggest_reference(request):
    doc_type = request.GET.get('doc_type', 'INCOMING')
    return JsonResponse({'reference': generate_reference_number(doc_type)})


# ──────────────────────────── CATEGORIES ─────────────────────────────

@login_required
def category_list(request):
    cats = Category.objects.annotate(cnt=Count('documents')).order_by('name')
    return render(request, 'documents/category_list.html', {'cats': cats})


@login_required
@registrar_required
def category_create(request):
    form = CategoryForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'ምድብ ተፈጠረ!')
        return redirect('category_list')
    return render(request, 'documents/category_form.html', {'form': form, 'action': 'ፍጠር'})


@login_required
@registrar_required
def category_update(request, pk):
    cat  = get_object_or_404(Category, pk=pk)
    form = CategoryForm(request.POST or None, instance=cat)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'ምድብ ተዘምኗል!')
        return redirect('category_list')
    return render(request, 'documents/category_form.html', {
        'form': form, 'cat': cat, 'action': 'አርትዕ'
    })


@login_required
@registrar_required
def category_delete(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        cat.delete()
        messages.warning(request, f'ምድብ "{cat}" ተሰርዟል።')
        return redirect('category_list')
    return render(request, 'documents/category_confirm_delete.html', {'cat': cat})


# ──────────────────────────── PROFILE ────────────────────────────────

@login_required
def profile_view(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        # Update user info
        request.user.first_name = request.POST.get('first_name', '').strip()
        request.user.last_name  = request.POST.get('last_name', '').strip()
        request.user.email      = request.POST.get('email', '').strip()
        request.user.save(update_fields=['first_name', 'last_name', 'email'])
        # Update profile
        profile.department = request.POST.get('department', '').strip()
        profile.phone      = request.POST.get('phone', '').strip()
        profile.save(update_fields=['department', 'phone'])
        messages.success(request, 'መረጃዎ ተዘምኗል!')
        return redirect('profile')

    my_docs     = Document.objects.filter(created_by=request.user).order_by('-created_at')[:5]
    my_assigned = Document.objects.filter(assigned_to=request.user).exclude(
                    status__in=['COMPLETED','ARCHIVED']).order_by('-created_at')[:5]
    my_total    = Document.objects.filter(created_by=request.user).count()
    return render(request, 'documents/profile.html', {
        'profile':     profile,
        'my_docs':     my_docs,
        'my_assigned': my_assigned,
        'my_total':    my_total,
    })


# ──────────────────────────── MY DOCUMENTS ───────────────────────────

@login_required
def my_documents(request):
    today = timezone.localdate()
    assigned = Document.objects.filter(
        assigned_to=request.user
    ).select_related('category').order_by('-created_at')

    created = Document.objects.filter(
        created_by=request.user
    ).select_related('category').order_by('-created_at')

    status_filter = request.GET.get('status', '')
    if status_filter:
        assigned = assigned.filter(status=status_filter)
        created  = created.filter(status=status_filter)

    return render(request, 'documents/my_documents.html', {
        'assigned':      assigned[:20],
        'created':       created[:20],
        'assigned_count': assigned.count(),
        'created_count':  created.count(),
        'status_filter':  status_filter,
        'status_choices': Document.STATUS_CHOICES,
        'today':          today,
    })


# ──────────────────────────── CREATE USER ────────────────────────────

@login_required
def create_user_view(request):
    """ADMIN-only: create a new user with role and department."""
    from .decorators import _get_role
    if _get_role(request.user) != 'ADMIN':
        messages.error(request, 'ይህን ተግባር ለማከናወን የአስተዳዳሪ ፈቃድ ያስፈልጋል።')
        return redirect('dashboard')

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        password   = request.POST.get('password', '').strip()
        password2  = request.POST.get('password2', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        role       = request.POST.get('role', 'VIEWER')
        department = request.POST.get('department', '').strip()
        phone      = request.POST.get('phone', '').strip()

        # Validation
        errors = []
        if not username:
            errors.append('Username ያስፈልጋል።')
        elif User.objects.filter(username=username).exists():
            errors.append(f'"{username}" አስቀድሞ ተመዝግቧል።')
        if not password:
            errors.append('Password ያስፈልጋል።')
        elif len(password) < 8:
            errors.append('Password ቢያንስ 8 ፊደላት ይኑሩት።')
        elif password != password2:
            errors.append('Passwords አይዛመዱም።')

        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'documents/create_user.html', {
                'role_choices': UserProfile.ROLE_CHOICES,
                'post': request.POST,
            })

        # Create user
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=first_name,
            last_name=last_name,
        )
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.role       = role
        profile.department = department
        profile.phone      = phone
        profile.save()

        messages.success(request, f'ተጠቃሚ "{username}" ({first_name} {last_name}) ተፈጠረ!')
        return redirect('user_management')

    return render(request, 'documents/create_user.html', {
        'role_choices': UserProfile.ROLE_CHOICES,
        'post': {},
    })


# ──────────────────────────── DOCUMENT FORWARD ───────────────────────

@login_required
@registrar_required
def document_forward(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    users = User.objects.filter(is_active=True).exclude(pk=request.user.pk).select_related('profile')

    if request.method == 'POST':
        to_user_id   = request.POST.get('to_user')
        to_department = request.POST.get('to_department', '').strip()
        note         = request.POST.get('note', '').strip()

        to_user = None
        if to_user_id:
            try:
                to_user = User.objects.get(pk=to_user_id)
            except User.DoesNotExist:
                pass

        from .models import DocumentForward
        fwd = DocumentForward.objects.create(
            document=doc,
            from_user=request.user,
            to_user=to_user,
            to_department=to_department,
            note=note,
        )

        # Log activity
        dest = to_user.get_full_name() or to_user.username if to_user else to_department
        log_activity(doc, request.user, 'ደብዳቤ ተላልፏል', f'ወደ {dest} ተላልፏል። {note}')

        # Notify recipient
        if to_user:
            Notification.objects.create(
                user=to_user,
                message=f'ደብዳቤ "{doc.reference_number} — {doc.title}" ወደዎ ተላልፏል። ማስታወሻ: {note or "—"}',
                link=f'/documents/{doc.pk}/',
            )
            # Send email if user has email
            if to_user.email:
                _send_forward_email(to_user, doc, request.user, note)

        messages.success(request, f'ደብዳቤ ወደ "{dest}" ተላልፏል!')
        return redirect('document_detail', pk=doc.pk)

    return render(request, 'documents/document_forward.html', {
        'doc': doc, 'users': users,
    })


def _send_forward_email(to_user, doc, from_user, note):
    """Send email notification when a document is forwarded."""
    try:
        from django.core.mail import send_mail
        subject = f'[EDMS] ደብዳቤ ተላልፏል — {doc.reference_number}'
        message = f"""
ሰላም {to_user.get_full_name() or to_user.username}،

{from_user.get_full_name() or from_user.username} ደብዳቤ ወደዎ አስተላልፏል።

ደብዳቤ: {doc.reference_number} — {doc.title}
ላካ: {from_user.get_full_name() or from_user.username}
ማስታወሻ: {note or '—'}

ለማየት: https://muleedms.pythonanywhere.com/documents/{doc.pk}/

ወልድያ ከተማ አስተዳደር — EDMS
        """
        send_mail(
            subject, message,
            settings.DEFAULT_FROM_EMAIL,
            [to_user.email],
            fail_silently=True,
        )
    except Exception:
        pass


# ──────────────────────────── INBOX (Forwarded to me) ────────────────

@login_required
def inbox(request):
    from .models import DocumentForward
    forwards = DocumentForward.objects.filter(
        to_user=request.user
    ).select_related('document', 'from_user').order_by('-forwarded_at')

    unread = forwards.filter(is_read=False).count()

    if request.GET.get('mark_read'):
        forwards.filter(is_read=False).update(is_read=True)
        return redirect('inbox')

    return render(request, 'documents/inbox.html', {
        'forwards': forwards,
        'unread':   unread,
    })


@login_required
@require_POST
def mark_forward_read(request, pk):
    from .models import DocumentForward
    fwd = get_object_or_404(DocumentForward, pk=pk, to_user=request.user)
    fwd.is_read = True
    fwd.save(update_fields=['is_read'])
    return redirect('document_detail', pk=fwd.document.pk)


# ──────────────────────────── OFFICIAL LETTER PDF ────────────────────

@login_required
def official_letter_pdf(request, pk):
    """Generate official Woldiya City Administration letter PDF."""
    doc = get_object_or_404(Document, pk=pk)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable, Image)
        from reportlab.lib.units import cm
        from io import BytesIO

        buf  = BytesIO()
        page = SimpleDocTemplate(buf, pagesize=A4,
                                  rightMargin=2.5*cm, leftMargin=2.5*cm,
                                  topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        brand  = colors.HexColor('#1B4F72')
        gold   = colors.HexColor('#C8960C')
        story  = []

        # ── Header with logo ──
        header_data = [[
            Paragraph(
                '<b>ወልድያ ከተማ አስተዳደር</b><br/>'
                '<font size="10">WOLDIA CITY ADMINISTRATION</font><br/>'
                '<font size="9" color="#1B4F72">ፈጠራ እና ቴክኖሎጂ ቡድን — EDMS</font>',
                ParagraphStyle('org', fontSize=14, textColor=brand, leading=18)
            ),
            Paragraph(
                f'<b>ቁጥር:</b> {doc.reference_number}<br/>'
                f'<b>ቀን:</b> {doc.created_at.strftime("%d/%m/%Y")}<br/>'
                f'<b>ዓይነት:</b> {doc.get_doc_type_display()}',
                ParagraphStyle('ref', fontSize=10, textColor=colors.HexColor('#374151'), leading=15)
            ),
        ]]

        # Try to add logo
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'woldiya_logo.png')
        if os.path.exists(logo_path):
            try:
                logo_img = Image(logo_path, width=2.5*cm, height=2.5*cm)
                header_data[0].insert(0, logo_img)
                col_widths = [2.8*cm, 10*cm, 4*cm]
            except Exception:
                col_widths = [12*cm, 4*cm]
        else:
            col_widths = [12*cm, 4*cm]

        t = Table(header_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN',  (-1,0), (-1,-1), 'RIGHT'),
        ]))
        story.append(t)
        story.append(HRFlowable(width='100%', color=gold, thickness=3, spaceAfter=16))

        # ── Document type banner ──
        type_style = ParagraphStyle('type', fontSize=14, fontName='Helvetica-Bold',
                                     textColor=colors.white, backColor=brand,
                                     alignment=1, spaceAfter=16, leading=20,
                                     leftIndent=-20, rightIndent=-20)
        story.append(Paragraph(f'  {doc.get_doc_type_display().upper()}  ', type_style))

        # ── Recipients ──
        recip_data = [
            [Paragraph('<b>ለ:</b>', styles['Normal']),
             Paragraph(doc.receiver, styles['Normal'])],
            [Paragraph('<b>ከ:</b>', styles['Normal']),
             Paragraph(doc.sender, styles['Normal'])],
            [Paragraph('<b>ጉዳዩ:</b>', styles['Normal']),
             Paragraph(f'<b>{doc.title}</b>', styles['Normal'])],
        ]
        if doc.due_date:
            recip_data.append([
                Paragraph('<b>የቀን ገደብ:</b>', styles['Normal']),
                Paragraph(str(doc.due_date), styles['Normal']),
            ])

        t2 = Table(recip_data, colWidths=[3.5*cm, 12.5*cm])
        t2.setStyle(TableStyle([
            ('FONTSIZE',  (0,0), (-1,-1), 11),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING',    (0,0), (-1,-1), 6),
            ('LINEBELOW', (0,0), (-1,-2), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        story.append(t2)
        story.append(Spacer(1, 16))
        story.append(HRFlowable(width='100%', color=colors.HexColor('#E2E8F0'), thickness=1))
        story.append(Spacer(1, 12))

        # ── Body / Description ──
        if doc.description:
            story.append(Paragraph('<b>ዝርዝር ይዘት / ማብራሪያ:</b>',
                                    ParagraphStyle('h', fontSize=11, fontName='Helvetica-Bold',
                                                   textColor=brand, spaceAfter=8)))
            story.append(Paragraph(doc.description.replace('\n', '<br/>'),
                                    ParagraphStyle('body', fontSize=11, leading=18,
                                                   spaceAfter=20)))

        # ── Status & Priority ──
        story.append(Spacer(1, 8))
        status_data = [[
            Paragraph(f'<b>ሁኔታ:</b> {doc.get_status_display()}', styles['Normal']),
            Paragraph(f'<b>ቅድሚያ:</b> {doc.get_priority_display()}', styles['Normal']),
            Paragraph(f'<b>ምድብ:</b> {doc.category or "—"}', styles['Normal']),
        ]]
        t3 = Table(status_data, colWidths=[5*cm, 5*cm, 6*cm])
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F0F4F8')),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('PADDING', (0,0), (-1,-1), 8),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(t3)
        story.append(Spacer(1, 40))

        # ── Signature area ──
        sig_data = [[
            Paragraph('___________________________<br/><font size="9">የሰነዱ አዘጋጅ ፊርማ</font>',
                       ParagraphStyle('sig', fontSize=10, alignment=1, leading=14)),
            Paragraph('___________________________<br/><font size="9">ኃላፊ ፊርማ</font>',
                       ParagraphStyle('sig', fontSize=10, alignment=1, leading=14)),
            Paragraph('___________________________<br/><font size="9">ቀን</font>',
                       ParagraphStyle('sig', fontSize=10, alignment=1, leading=14)),
        ]]
        t4 = Table(sig_data, colWidths=[5.5*cm, 5.5*cm, 5*cm])
        t4.setStyle(TableStyle([('TOPPADDING', (0,0), (-1,-1), 20)]))
        story.append(t4)

        # ── Footer ──
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width='100%', color=gold, thickness=2))
        story.append(Paragraph(
            '<font size="8" color="#94A3B8">ወልድያ ከተማ አስተዳደር — ፈጠራ እና ቴክኖሎጂ ቡድን | EDMS | '
            f'ህትመት ቀን: {timezone.localdate()} | ቁጥር: {doc.reference_number}</font>',
            ParagraphStyle('footer', fontSize=8, alignment=1, spaceAfter=0, leading=12)
        ))

        page.build(story)
        buf.seek(0)
        response = HttpResponse(buf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = (
            f'inline; filename="letter_{doc.reference_number}.pdf"'
        )
        return response

    except ImportError:
        messages.error(request, 'reportlab አልተጫነም። pip install reportlab')
        return redirect('document_detail', pk=pk)
    except Exception as e:
        messages.error(request, f'PDF ስህተት: {e}')
        return redirect('document_detail', pk=pk)
