from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Count, F
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.conf import settings
import json, datetime, csv, os

from .models import (Document, Category, Notification,
                     InnovationProject, Training, UserProfile, ActivityLog)
from .forms import DocumentForm, DocumentSearchForm, CategoryForm
from .activity import log_activity, notify_users
from .decorators import registrar_required
from .utils import generate_reference_number, generate_qr_code


# ──────────────────────────── AUTH ───────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        login(request, user)
        messages.success(request, f'እንኳን ደህና መጡ, {user.get_full_name() or user.username}!')
        return redirect(request.GET.get('next', 'dashboard'))
    return render(request, 'documents/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


# ──────────────────────────── DASHBOARD ──────────────────────────────

@login_required
def dashboard(request):
    today      = timezone.localdate()
    week_start = today - datetime.timedelta(days=today.weekday())
    docs       = Document.objects.all()

    total       = docs.count()
    incoming    = docs.filter(doc_type='INCOMING').count()
    outgoing    = docs.filter(doc_type='OUTGOING').count()
    internal    = docs.filter(doc_type='INTERNAL').count()
    pending     = docs.filter(status='PENDING').count()
    in_progress = docs.filter(status='IN_PROGRESS').count()
    completed   = docs.filter(status='COMPLETED').count()
    urgent      = docs.filter(priority='URGENT').count()
    overdue     = docs.filter(due_date__lt=today).exclude(
                    status__in=['COMPLETED', 'ARCHIVED']).count()
    today_docs  = docs.filter(created_at__date=today).count()
    week_docs   = docs.filter(created_at__date__gte=week_start).count()
    my_docs     = docs.filter(assigned_to=request.user).exclude(
                    status__in=['COMPLETED', 'ARCHIVED']).count()

    recent_docs   = docs.select_related('category', 'created_by').order_by('-created_at')[:8]
    deadline_soon = docs.filter(
        due_date__gte=today,
        due_date__lte=today + datetime.timedelta(days=7)
    ).exclude(status__in=['COMPLETED', 'ARCHIVED']).order_by('due_date')[:6]

    chart_labels, chart_data = [], []
    for i in range(5, -1, -1):
        d = today - datetime.timedelta(days=30 * i)
        chart_labels.append(d.strftime('%b %Y'))
        chart_data.append(docs.filter(created_at__year=d.year, created_at__month=d.month).count())

    type_data    = {'ገቢ ደብዳቤ': incoming, 'ወጪ ደብዳቤ': outgoing, 'የውስጥ ማስታወሻ': internal}
    status_labels = [s[1] for s in Document.STATUS_CHOICES]
    status_values = [docs.filter(status=s[0]).count() for s in Document.STATUS_CHOICES]

    return render(request, 'documents/dashboard.html', {
        'total': total, 'incoming': incoming, 'outgoing': outgoing, 'internal': internal,
        'pending': pending, 'in_progress': in_progress, 'completed': completed,
        'urgent': urgent, 'overdue': overdue,
        'today_docs': today_docs, 'week_docs': week_docs, 'my_docs': my_docs,
        'recent_docs': recent_docs, 'deadline_soon': deadline_soon,
        'chart_labels':  json.dumps(chart_labels),
        'chart_data':    json.dumps(chart_data),
        'type_labels':   json.dumps(list(type_data.keys())),
        'type_values':   json.dumps(list(type_data.values())),
        'status_labels': json.dumps(status_labels),
        'status_values': json.dumps(status_values),
        'today': today,
        'completion_pct': int((completed / total * 100) if total else 0),
    })


# ──────────────────────────── DOCUMENT LIST ──────────────────────────

@login_required
def document_list(request):
    form = DocumentSearchForm(request.GET or None)
    docs = Document.objects.select_related('category', 'created_by').all()

    if form.is_valid():
        q = form.cleaned_data.get('query')
        if q:
            docs = docs.filter(
                Q(title__icontains=q) | Q(reference_number__icontains=q) |
                Q(sender__icontains=q) | Q(receiver__icontains=q) |
                Q(description__icontains=q)
            )
        for field in ('doc_type', 'status', 'priority', 'category'):
            val = form.cleaned_data.get(field)
            if val:
                docs = docs.filter(**{field: val})
        if form.cleaned_data.get('date_from'):
            docs = docs.filter(created_at__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            docs = docs.filter(created_at__date__lte=form.cleaned_data['date_to'])

    if request.GET.get('export') == 'csv':
        return _export_csv(docs)

    # Sorting
    sort = request.GET.get('sort', '-created_at')
    allowed_sorts = ['reference_number', '-reference_number', 'created_at', '-created_at',
                     'due_date', '-due_date', 'title', '-title']
    if sort not in allowed_sorts:
        sort = '-created_at'
    docs = docs.order_by(sort)

    paginator = Paginator(docs, 15)
    page_obj  = paginator.get_page(request.GET.get('page'))
    today     = timezone.localdate()
    return render(request, 'documents/document_list.html', {
        'form': form, 'page_obj': page_obj,
        'total_results': paginator.count,
        'today_date': today,
        'today_plus7': today + datetime.timedelta(days=7),
    })


def _export_csv(queryset):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="documents.csv"'
    response.write('\ufeff')
    w = csv.writer(response)
    w.writerow(['ቁጥር', 'ርዕስ', 'ዓይነት', 'ሁኔታ', 'ቅድሚያ', 'ላኪ', 'ተቀባይ', 'ቀን'])
    for d in queryset:
        w.writerow([
            d.reference_number, d.title, d.get_doc_type_display(),
            d.get_status_display(), d.get_priority_display(),
            d.sender, d.receiver, d.created_at.strftime('%d/%m/%Y'),
        ])
    return response


# ──────────────────────────── DOCUMENT DETAIL ────────────────────────

@login_required
def document_detail(request, pk):
    doc        = get_object_or_404(Document, pk=pk)
    activities = doc.activities.select_related('user').order_by('-timestamp')[:20]
    today      = timezone.localdate()
    is_overdue = (doc.due_date and doc.due_date < today
                  and doc.status not in ['COMPLETED', 'ARCHIVED'])
    Document.objects.filter(pk=pk).update(view_count=F('view_count') + 1)
    return render(request, 'documents/document_detail.html', {
        'doc': doc, 'activities': activities, 'is_overdue': is_overdue,
    })


# ──────────────────────────── CREATE ─────────────────────────────────

@login_required
@registrar_required
def document_create(request):
    form = DocumentForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        doc = form.save(commit=False)
        doc.created_by = request.user
        doc.save()
        log_activity(doc, request.user, 'ደብዳቤ ተመዝግቧል', f'"{doc.title}" ተመዝግቧል።')
        notify_users(
            f'አዲስ ደብዳቤ "{doc.reference_number} — {doc.title}" ተመዝግቧል።',
            link=f'/documents/{doc.pk}/', exclude_user=request.user,
        )
        messages.success(request, f'ደብዳቤ "{doc.title}" ተመዝግቧል!')
        return redirect('document_detail', pk=doc.pk)
    return render(request, 'documents/document_form.html', {'form': form, 'action': 'ምዝገባ'})


# ──────────────────────────── UPDATE ─────────────────────────────────

@login_required
@registrar_required
def document_update(request, pk):
    doc        = get_object_or_404(Document, pk=pk)
    old_status = doc.status
    form       = DocumentForm(request.POST or None, request.FILES or None, instance=doc)
    if request.method == 'POST' and form.is_valid():
        updated = form.save()
        if old_status != updated.status:
            log_activity(doc, request.user, 'ሁኔታ ተቀይሯል',
                         f'"{doc.get_status_display()}" → "{updated.get_status_display()}"')
        else:
            log_activity(doc, request.user, 'ደብዳቤ ተዘምኗል')
        messages.success(request, 'ደብዳቤ ተዘምኗል!')
        return redirect('document_detail', pk=doc.pk)
    return render(request, 'documents/document_form.html', {
        'form': form, 'doc': doc, 'action': 'ማስተካከያ'
    })


# ──────────────────────────── DELETE ─────────────────────────────────

@login_required
@registrar_required
def document_delete(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
        title = doc.title
        doc.delete()
        messages.warning(request, f'ደብዳቤ "{title}" ተሰርዟል።')
        return redirect('document_list')
    return render(request, 'documents/document_confirm_delete.html', {'doc': doc})


# ──────────────────────────── QUICK STATUS ───────────────────────────

@login_required
@registrar_required
@require_POST
def update_status(request, pk):
    doc        = get_object_or_404(Document, pk=pk)
    new_status = request.POST.get('status')
    valid      = [s[0] for s in Document.STATUS_CHOICES]
    if new_status in valid:
        old = doc.get_status_display()
        doc.status = new_status
        doc.save(update_fields=['status', 'updated_at'])
        log_activity(doc, request.user, 'ሁኔታ ተቀይሯል',
                     f'"{old}" → "{doc.get_status_display()}"')
        return JsonResponse({'success': True, 'new_status': doc.get_status_display()})
    return JsonResponse({'success': False}, status=400)


# ──────────────────────────── DOWNLOAD ───────────────────────────────

@login_required
def document_download(request, pk):
    from django.http import FileResponse, Http404
    doc = get_object_or_404(Document, pk=pk)
    if not doc.file:
        messages.error(request, 'ፋይሉ አልተገኘም።')
        return redirect('document_detail', pk=pk)
    try:
        file_path = doc.file.path
        if not os.path.exists(file_path):
            raise Http404
        Document.objects.filter(pk=pk).update(download_count=F('download_count') + 1)
        log_activity(doc, request.user, 'ፋይል ወረደ',
                     f'"{os.path.basename(file_path)}" ወረደ።')
        response = FileResponse(open(file_path, 'rb'), as_attachment=True)
        response['Content-Disposition'] = (
            f'attachment; filename="{os.path.basename(file_path)}"'
        )
        return response
    except Exception:
        messages.error(request, 'ፋይሉ ከሰርቨሩ ላይ አልተገኘም።')
        return redirect('document_detail', pk=pk)


# ──────────────────────────── PRINT & QR ─────────────────────────────

@login_required
def document_print(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    return render(request, 'documents/document_print.html', {'doc': doc})


@login_required
def document_qr(request, pk):
    """Return the QR code PNG — with Woldiya logo embedded in the centre."""
    doc      = get_object_or_404(Document, pk=pk)
    scheme   = 'https' if request.is_secure() else 'http'
    doc_url  = f"{scheme}://{request.get_host()}/documents/{doc.pk}/"
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'woldiya_logo.png')
    png_bytes = generate_qr_code(doc_url, logo_path=logo_path)
    return HttpResponse(png_bytes, content_type='image/png')


# ──────────────────────────── OVERDUE ────────────────────────────────

@login_required
def overdue_documents(request):
    today = timezone.localdate()
    docs  = Document.objects.filter(due_date__lt=today).exclude(
                status__in=['COMPLETED', 'ARCHIVED']).order_by('due_date')
    return render(request, 'documents/overdue.html', {'docs': docs, 'today': today})


# ──────────────────────────── NOTIFICATIONS ──────────────────────────

@login_required
def notifications_view(request):
    # Mark all as read on visit
    if request.GET.get('mark_all'):
        request.user.notifications.filter(is_read=False).update(is_read=True)
        messages.success(request, 'ሁሉም ማሳወቂያዎች ተነቡ።')
        return redirect('notifications')

    notifs = request.user.notifications.order_by('-created_at')[:100]
    # Group by date
    from itertools import groupby
    from django.utils.timezone import localdate
    grouped = {}
    for n in notifs:
        day = localdate(n.created_at)
        grouped.setdefault(day, []).append(n)

    return render(request, 'documents/notifications.html', {
        'notifs': notifs,
        'grouped': grouped,
        'unread_count': request.user.notifications.filter(is_read=False).count(),
        'today': timezone.localdate(),
    })


@login_required
@require_POST
def mark_notification_read(request, pk):
    n = get_object_or_404(Notification, pk=pk, user=request.user)
    n.is_read = True
    n.save(update_fields=['is_read'])
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    return redirect(n.link or 'notifications')


@login_required
@require_POST
def mark_all_notifications_read(request):
    request.user.notifications.filter(is_read=False).update(is_read=True)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    messages.success(request, 'ሁሉም ማሳወቂያዎች ተነቡ።')
    return redirect('notifications')


# ──────────────────────────── AUDIT TRAIL ────────────────────────────

@login_required
def audit_trail(request):
    logs = ActivityLog.objects.select_related('document', 'user').order_by('-timestamp')

    # Filters
    user_q   = request.GET.get('user', '').strip()
    action_q = request.GET.get('action', '').strip()
    date_f   = request.GET.get('date_from', '').strip()
    date_t   = request.GET.get('date_to', '').strip()

    if user_q:
        logs = logs.filter(
            Q(user__username__icontains=user_q) |
            Q(user__first_name__icontains=user_q) |
            Q(user__last_name__icontains=user_q)
        )
    if action_q:
        logs = logs.filter(action__icontains=action_q)
    if date_f:
        try:
            logs = logs.filter(timestamp__date__gte=datetime.date.fromisoformat(date_f))
        except ValueError:
            pass
    if date_t:
        try:
            logs = logs.filter(timestamp__date__lte=datetime.date.fromisoformat(date_t))
        except ValueError:
            pass

    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="audit_trail.csv"'
        response.write('\ufeff')
        w = csv.writer(response)
        w.writerow(['ቀን', 'ተጠቃሚ', 'ደብዳቤ', 'እርምጃ', 'ዝርዝር'])
        for log in logs:
            w.writerow([
                log.timestamp.strftime('%d/%m/%Y %H:%M'),
                log.user.get_full_name() or log.user.username if log.user else '—',
                log.document.reference_number,
                log.action,
                log.detail or '—',
            ])
        return response

    paginator = Paginator(logs, 30)
    page_obj  = paginator.get_page(request.GET.get('page'))
    users     = User.objects.filter(is_active=True).order_by('username')
    return render(request, 'documents/audit_trail.html', {
        'page_obj': page_obj, 'users': users,
        'filter_user': user_q, 'filter_action': action_q,
        'filter_date_from': date_f, 'filter_date_to': date_t,
    })


# ──────────────────────────── REPORTS ────────────────────────────────

@login_required
def reports(request):
    today = timezone.localdate()
    docs  = Document.objects.all()

    by_type     = {t: docs.filter(doc_type=t).count() for t, _ in Document.DOCUMENT_TYPES}
    by_status   = {s: docs.filter(status=s).count()   for s, _ in Document.STATUS_CHOICES}
    by_priority = {p: docs.filter(priority=p).count() for p, _ in Document.PRIORITY_CHOICES}
    by_category = Category.objects.annotate(cnt=Count('documents')).order_by('-cnt')

    monthly = []
    for i in range(11, -1, -1):
        d = today.replace(day=1) - datetime.timedelta(days=30 * i)
        monthly.append({
            'label': d.strftime('%b %Y'),
            'count': docs.filter(created_at__year=d.year, created_at__month=d.month).count(),
        })

    if request.GET.get('export') == 'excel':
        return _export_excel_report(docs, by_type, by_status, by_priority, monthly, today)
    if request.GET.get('export') == 'pdf':
        return _export_pdf_report(docs, by_type, by_status, by_priority, monthly, today)

    return render(request, 'documents/reports.html', {
        'total': docs.count(),
        'by_type': by_type, 'by_status': by_status,
        'by_priority': by_priority, 'by_category': by_category,
        'monthly': monthly, 'today': today,
    })


def _export_excel_report(docs, by_type, by_status, by_priority, monthly, today):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl አልተጫነም። pip install openpyxl', status=500)

    wb = openpyxl.Workbook()
    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='1B4F72')
    center    = Alignment(horizontal='center')

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = 'ማጠቃለያ'
    ws.append(['ወልድያ ከተማ አስተዳደር — EDMS ሪፖርት'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'ዛሬ: {today}'])
    ws.append([])
    ws.append(['ጠቅላላ', 'ዓይነት', 'ብዛት'])
    for cell in ws[4]:
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
    ws.append(['ጠቅላላ ደብዳቤ', '', docs.count()])
    for k, v in by_type.items():
        ws.append([k, '', v])
    ws.append([])
    ws.append(['ሁኔታ', '', 'ብዛት'])
    for k, v in by_status.items():
        ws.append([k, '', v])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['C'].width = 12

    # ── Sheet 2: Monthly ──
    ws2 = wb.create_sheet('ወርሃዊ')
    ws2.append(['ወር', 'ደብዳቤ ብዛት'])
    for cell in ws2[1]:
        cell.font = hdr_font; cell.fill = hdr_fill
    for row in monthly:
        ws2.append([row['label'], row['count']])
    ws2.column_dimensions['A'].width = 18

    # ── Sheet 3: All documents ──
    ws3 = wb.create_sheet('ሁሉም ደብዳቤዎች')
    ws3.append(['ቁጥር', 'ርዕስ', 'ዓይነት', 'ሁኔታ', 'ቅድሚያ', 'ላኪ', 'ተቀባይ', 'ቀን'])
    for cell in ws3[1]:
        cell.font = hdr_font; cell.fill = hdr_fill
    for d in docs.order_by('-created_at'):
        ws3.append([
            d.reference_number, d.title, d.get_doc_type_display(),
            d.get_status_display(), d.get_priority_display(),
            d.sender, d.receiver, d.created_at.strftime('%d/%m/%Y'),
        ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws3.column_dimensions[col].width = 20

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="edms_report_{today}.xlsx"'
    return response


def _export_pdf_report(docs, by_type, by_status, by_priority, monthly, today):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return HttpResponse('reportlab አልተጫነም። pip install reportlab', status=500)

    from io import BytesIO
    buf  = BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                              rightMargin=2*cm, leftMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    brand  = colors.HexColor('#1B4F72')
    gold   = colors.HexColor('#C8960C')
    story  = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                  textColor=brand, fontSize=18, spaceAfter=6)
    story.append(Paragraph('ወልድያ ከተማ አስተዳደር — EDMS ሪፖርት', title_style))
    story.append(Paragraph(f'ዛሬ: {today}', styles['Normal']))
    story.append(HRFlowable(width='100%', color=gold, thickness=2, spaceAfter=12))

    # Summary table
    summary_data = [['ምድብ', 'ብዛት']]
    summary_data.append(['ጠቅላላ ደብዳቤ', str(docs.count())])
    for k, v in by_type.items():
        summary_data.append([k, str(v)])
    for k, v in by_status.items():
        summary_data.append([k, str(v)])

    t = Table(summary_data, colWidths=[12*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), brand),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN',      (1, 0), (1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('PADDING',    (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    # Monthly table
    story.append(Paragraph('ወርሃዊ ምዝገባ', styles['Heading2']))
    month_data = [['ወር', 'ደብዳቤ']]
    for row in monthly:
        month_data.append([row['label'], str(row['count'])])
    t2 = Table(month_data, colWidths=[12*cm, 4*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), brand),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN',      (1, 0), (1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('FONTSIZE',   (0, 0), (-1, -1), 10),
        ('PADDING',    (0, 0), (-1, -1), 6),
    ]))
    story.append(t2)

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="edms_report_{today}.pdf"'
    return response


# ──────────────────────────── CAPACITY BUILDING ──────────────────────

@login_required
def capacity_building(request):
    projects  = InnovationProject.objects.select_related('lead').order_by('-created_at')
    trainings = Training.objects.prefetch_related('participants').order_by('-date')
    return render(request, 'documents/capacity_building.html', {
        'projects': projects, 'trainings': trainings,
    })


# ──────────────────────────── USER MANAGEMENT ────────────────────────

@login_required
def user_management(request):
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'ፈቃድ የለዎትም።')
        return redirect('dashboard')
    users = User.objects.select_related('profile').filter(is_active=True).order_by('username')
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        role    = request.POST.get('role')
        dept    = request.POST.get('department', '').strip()
        try:
            profile = UserProfile.objects.get(user_id=user_id)
            if role in dict(UserProfile.ROLE_CHOICES):
                profile.role       = role
                profile.department = dept
                profile.save()
                messages.success(request, 'ተጠቃሚ ተዘምኗል።')
        except UserProfile.DoesNotExist:
            messages.error(request, 'ተጠቃሚ አልተገኘም።')
        return redirect('user_management')
    return render(request, 'documents/user_management.html', {
        'users': users, 'role_choices': UserProfile.ROLE_CHOICES,
    })


# ──────────────────────────── API HELPERS ────────────────────────────

@login_required
def suggest_reference(request):
    doc_type = request.GET.get('doc_type', 'INCOMING')
    return JsonResponse({'reference': generate_reference_number(doc_type)})


# ──────────────────────────── CATEGORIES ─────────────────────────────

@login_required
def category_list(request):
    cats = Category.objects.annotate(cnt=Count('documents')).order_by('name')
    return render(request, 'documents/category_list.html', {'cats': cats})


@login_required
@registrar_required
def category_create(request):
    form = CategoryForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'ምድብ ተፈጠረ!')
        return redirect('category_list')
    return render(request, 'documents/category_form.html', {'form': form, 'action': 'ፍጠር'})


@login_required
@registrar_required
def category_update(request, pk):
    cat  = get_object_or_404(Category, pk=pk)
    form = CategoryForm(request.POST or None, instance=cat)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'ምድብ ተዘምኗል!')
        return redirect('category_list')
    return render(request, 'documents/category_form.html', {
        'form': form, 'cat': cat, 'action': 'አርትዕ'
    })


@login_required
@registrar_required
def category_delete(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        cat.delete()
        messages.warning(request, f'ምድብ "{cat}" ተሰርዟል።')
        return redirect('category_list')
    return render(request, 'documents/category_confirm_delete.html', {'cat': cat})


# ──────────────────────────── PROFILE ────────────────────────────────

@login_required
def profile_view(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        # Update user info
        request.user.first_name = request.POST.get('first_name', '').strip()
        request.user.last_name  = request.POST.get('last_name', '').strip()
        request.user.email      = request.POST.get('email', '').strip()
        request.user.save(update_fields=['first_name', 'last_name', 'email'])
        # Update profile
        profile.department = request.POST.get('department', '').strip()
        profile.phone      = request.POST.get('phone', '').strip()
        profile.save(update_fields=['department', 'phone'])
        messages.success(request, 'መረጃዎ ተዘምኗል!')
        return redirect('profile')

    my_docs     = Document.objects.filter(created_by=request.user).order_by('-created_at')[:5]
    my_assigned = Document.objects.filter(assigned_to=request.user).exclude(
                    status__in=['COMPLETED','ARCHIVED']).order_by('-created_at')[:5]
    my_total    = Document.objects.filter(created_by=request.user).count()
    return render(request, 'documents/profile.html', {
        'profile':     profile,
        'my_docs':     my_docs,
        'my_assigned': my_assigned,
        'my_total':    my_total,
    })


# ──────────────────────────── MY DOCUMENTS ───────────────────────────

@login_required
def my_documents(request):
    today = timezone.localdate()
    assigned = Document.objects.filter(
        assigned_to=request.user
    ).select_related('category').order_by('-created_at')

    created = Document.objects.filter(
        created_by=request.user
    ).select_related('category').order_by('-created_at')

    status_filter = request.GET.get('status', '')
    if status_filter:
        assigned = assigned.filter(status=status_filter)
        created  = created.filter(status=status_filter)

    return render(request, 'documents/my_documents.html', {
        'assigned':      assigned[:20],
        'created':       created[:20],
        'assigned_count': assigned.count(),
        'created_count':  created.count(),
        'status_filter':  status_filter,
        'status_choices': Document.STATUS_CHOICES,
        'today':          today,
    })
